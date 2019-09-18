# -*- coding: utf-8 -*-
# from openpyxl import Workbook
# from openpyxl.styles import colors
# from openpyxl.styles import Font

# wb = Workbook()
# ws = wb.active

# a1 = ws['A1']
# d4 = ws['D4']
# ft = Font(color=colors.RED)  # color="FFBB00"，颜色编码也可以设定颜色
# a1.font = ft
# d4.font = ft

# # If you want to change the color of a Font, you need to reassign it::
# #italic 倾斜字体
# a1.font = Font(color=colors.RED, italic=True) # the change only affects A1
# a1.value = "abc"

# # Save the file
# wb.save("sample.xlsx")


# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import NamedStyle, Font, Border, Side,PatternFill

wb = Workbook()
ws = wb.active

# highlight = NamedStyle(name="highlight")
# highlight.font = Font(bold=True, size=20,color= "ff0100")
# highlight.fill = PatternFill("solid", fgColor="455921")#背景填充
# bd = Side(style='thick', color="000000")
# highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)

# print( dir(ws["A1"]))
ws["B3"].fill =PatternFill("solid", fgColor="455921")#背景填充
# ws["3C"].fill =PatternFill("solid", fgColor="455921")#背景填充
# ws.cell(1,2).style = highlight
# Save the file
wb.save("sample.xlsx")


from PIL import Image
########获取图片指定像素点的像素
def getPngPix(pngPath = "aa.jpg",pixelX = 1,pixelY = 1):
    img_src = Image.open(pngPath)
    print(img_src.size)
    img_src = img_src.convert('RGB')
    str_strlist = img_src.load()
    data = str_strlist[pixelX,pixelY]
    img_src.close()
    dat = data[0]+data[1]*256 + data[1]*65536

    return str(hex(dat)).split("0x")[1]
 
 
print(getPngPix())