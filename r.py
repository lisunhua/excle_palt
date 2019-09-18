# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import NamedStyle, Font, Border, Side,PatternFill
from PIL import Image

MAX_SCALA =26  

def datToStr_u(d):
    st="ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    dt=""
    while(d>=MAX_SCALA):    
        idx= int(d%MAX_SCALA)
        dt+=st[idx]
        d=int(d/MAX_SCALA)-1
    dt=st[d]+dt

    return dt


def full_execl(ws,x,y,strlist):
    for i in range(0,x-1):
        for k in range(1,y-1):
            # k=1
            data = str_strlist[i,k]
            dat = data[0]+data[1]*256 + data[1]*65536
            lod =datToStr_u(i)+str(k)
            col_str = str(hex(dat)).split("0x")[1]

            # if k == 1:
            #     print(str(i)+":>"+lod+":->"+col_str)
            #     pass
            while len(col_str)<6:
                col_str = "0"+ col_str
            ws[lod].fill = PatternFill("solid", fgColor=col_str)

wb = Workbook()
ws = wb.active
img_src  = Image.open("bb.jpg")
img_src  = img_src.convert('RGB')
img_size = img_src.size
print(img_size)
str_strlist = img_src.load()
full_execl(ws,img_size[0],img_size[1],str_strlist)
wb.save("sample.xlsx")
img_src.close()

